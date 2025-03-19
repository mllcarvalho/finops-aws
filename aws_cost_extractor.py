import boto3
import datetime
import pandas as pd
import os
import sys
import json
import webbrowser
import time
import re
from botocore.exceptions import ClientError
from dateutil.relativedelta import relativedelta
import argparse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.formatting.rule import ColorScaleRule
import locale

# Configurar localidade para formatação de números
try:
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_ALL, 'Portuguese_Brazil.1252')
    except:
        pass  # Fallback para configuração padrão


class AWSSSOAuth:
    """
    Classe para autenticação via AWS SSO e acesso às contas.
    """
    def __init__(self, start_url, region="us-east-1", sso_region=None):
        """
        Inicializa o autenticador SSO.
        
        Args:
            start_url: URL do portal SSO (ex: itaulzprod.awsapps.com/start)
            region: Região principal para API calls
            sso_region: Região do SSO (se diferente da principal)
        """
        self.start_url = start_url if start_url.startswith('https://') else f'https://{start_url}'
        self.region = region
        self.sso_region = sso_region if sso_region else region
        self.session = boto3.Session(region_name=self.region)
        self.sso_oidc_client = self.session.client('sso-oidc', region_name=self.sso_region)
        self.sso_client = self.session.client('sso', region_name=self.sso_region)
        self.token_cache_file = os.path.expanduser('~/.aws/sso/cache/token.json')
        self.accounts_cache_file = os.path.expanduser('~/.aws/sso/cache/accounts.json')
        
        # Criar diretório para cache se não existir
        os.makedirs(os.path.dirname(self.token_cache_file), exist_ok=True)
        
    def _get_cached_token(self):
        """
        Obtém token SSO do cache se válido.
        
        Returns:
            dict: Token de acesso ou None
        """
        if os.path.exists(self.token_cache_file):
            try:
                with open(self.token_cache_file, 'r') as f:
                    cache = json.load(f)
                    if 'expiresAt' in cache and 'accessToken' in cache:
                        expires_at = datetime.datetime.fromisoformat(cache['expiresAt'].replace('Z', '+00:00'))
                        if expires_at > datetime.datetime.now(datetime.timezone.utc):
                            return cache['accessToken']
            except Exception as e:
                print(f"Erro ao ler cache de token: {str(e)}")
        return None
        
    def _register_client(self):
        """
        Registra um cliente para o processo de autenticação.
        
        Returns:
            tuple: (client_id, client_secret)
        """
        response = self.sso_oidc_client.register_client(
            clientName='aws-cost-extractor',
            clientType='public'
        )
        return response['clientId'], response['clientSecret']
        
    def _start_device_authorization(self, client_id, client_secret):
        """
        Inicia o fluxo de autorização do dispositivo.
        
        Args:
            client_id: ID do cliente registrado
            client_secret: Secret do cliente registrado
            
        Returns:
            dict: Informações de autorização incluindo o URL para abrir no navegador
        """
        response = self.sso_oidc_client.start_device_authorization(
            clientId=client_id,
            clientSecret=client_secret,
            startUrl=self.start_url
        )
        return response
        
    def _open_browser_for_auth(self, verification_uri_complete):
        """
        Abre o navegador para autenticação.
        
        Args:
            verification_uri_complete: URL completo para autenticação
        """
        print(f"\nAbrindo navegador para autenticação AWS SSO...")
        print(f"Se o navegador não abrir automaticamente, acesse manualmente: {verification_uri_complete}")
        
        try:
            webbrowser.open(verification_uri_complete)
        except Exception as e:
            print(f"Erro ao abrir navegador: {str(e)}")
            print(f"Por favor, abra manualmente: {verification_uri_complete}")
        
        print("\nAguardando autenticação... (autorize no navegador e volte aqui)")
        
    def _get_token(self, client_id, client_secret, device_code):
        """
        Obtém o token após aprovação no navegador.
        
        Args:
            client_id: ID do cliente
            client_secret: Secret do cliente
            device_code: Código do dispositivo
            
        Returns:
            str: Token de acesso
        """
        # Tentar obter o token a cada 5 segundos até a autorização ou timeout
        max_tries = 60  # 5 minutos
        tries = 0
        
        while tries < max_tries:
            try:
                response = self.sso_oidc_client.create_token(
                    clientId=client_id,
                    clientSecret=client_secret,
                    grantType='urn:ietf:params:oauth:grant-type:device_code',
                    deviceCode=device_code
                )
                
                # Salvar token no cache
                cache = {
                    'accessToken': response['accessToken'],
                    'expiresAt': (datetime.datetime.now(datetime.timezone.utc) + 
                                 datetime.timedelta(seconds=response['expiresIn'])).isoformat()
                }
                
                with open(self.token_cache_file, 'w') as f:
                    json.dump(cache, f)
                    
                print("Autenticação bem-sucedida!")
                return response['accessToken']
                
            except self.sso_oidc_client.exceptions.AuthorizationPendingException:
                # Ainda aguardando aprovação
                time.sleep(5)
                tries += 1
            except Exception as e:
                print(f"Erro ao obter token: {str(e)}")
                return None
                
        print("Timeout na autenticação. Por favor, tente novamente.")
        return None
        
    def authenticate(self):
        """
        Executa o fluxo completo de autenticação.
        
        Returns:
            str: Token de acesso
        """
        # Verificar cache primeiro
        token = self._get_cached_token()
        if token:
            print("Usando token SSO em cache.")
            return token
            
        # Iniciar novo fluxo de autenticação
        print("Iniciando fluxo de autenticação AWS SSO...")
        client_id, client_secret = self._register_client()
        
        auth_response = self._start_device_authorization(client_id, client_secret)
        self._open_browser_for_auth(auth_response['verificationUriComplete'])
        
        return self._get_token(client_id, client_secret, auth_response['deviceCode'])
        
    def list_available_accounts(self, access_token):
        """
        Lista todas as contas disponíveis para o usuário.
        
        Args:
            access_token: Token de acesso SSO
            
        Returns:
            list: Lista de dicionários com informações das contas
        """
        # Verificar cache primeiro
        if os.path.exists(self.accounts_cache_file):
            try:
                with open(self.accounts_cache_file, 'r') as f:
                    cache = json.load(f)
                    if 'expiresAt' in cache and 'accounts' in cache:
                        expires_at = datetime.datetime.fromisoformat(cache['expiresAt'].replace('Z', '+00:00'))
                        if expires_at > datetime.datetime.now(datetime.timezone.utc):
                            print("Usando cache de contas disponíveis.")
                            return cache['accounts']
            except Exception as e:
                print(f"Erro ao ler cache de contas: {str(e)}")
        
        accounts = []
        try:
            paginator = self.sso_client.get_paginator('list_accounts')
            
            for page in paginator.paginate(accessToken=access_token):
                accounts.extend(page['accountList'])
                
            # Salvar no cache
            cache = {
                'accounts': accounts,
                'expiresAt': (datetime.datetime.now(datetime.timezone.utc) + 
                             datetime.timedelta(hours=24)).isoformat()
            }
            
            with open(self.accounts_cache_file, 'w') as f:
                json.dump(cache, f)
                
            return accounts
        except Exception as e:
            print(f"Erro ao listar contas: {str(e)}")
            return []
            
    def get_account_roles(self, access_token, account_id):
        """
        Obtém todas as funções disponíveis para uma conta.
        
        Args:
            access_token: Token de acesso SSO
            account_id: ID da conta AWS
            
        Returns:
            list: Lista de funções disponíveis
        """
        try:
            response = self.sso_client.list_account_roles(
                accessToken=access_token,
                accountId=account_id
            )
            return response['roleList']
        except Exception as e:
            print(f"Erro ao listar funções para conta {account_id}: {str(e)}")
            return []
            
    def get_credentials(self, access_token, account_id, role_name):
        """
        Obtém credenciais temporárias para uma conta/função.
        
        Args:
            access_token: Token de acesso SSO
            account_id: ID da conta AWS
            role_name: Nome da função a assumir
            
        Returns:
            boto3.Session: Sessão com as credenciais temporárias
        """
        try:
            response = self.sso_client.get_role_credentials(
                accessToken=access_token,
                accountId=account_id,
                roleName=role_name
            )
            
            creds = response['roleCredentials']
            
            return boto3.Session(
                aws_access_key_id=creds['accessKeyId'],
                aws_secret_access_key=creds['secretAccessKey'],
                aws_session_token=creds['sessionToken'],
                region_name=self.region
            )
        except Exception as e:
            print(f"Erro ao obter credenciais para conta {account_id}, função {role_name}: {str(e)}")
            return None


class AWSCostExtractor:
    def __init__(self, sso_auth, access_token):
        """
        Inicializa o extrator de custos AWS.
        
        Args:
            sso_auth: Instância de AWSSSOAuth
            access_token: Token de acesso SSO
        """
        self.sso_auth = sso_auth
        self.access_token = access_token
        self.account_data = []

    def get_cost_data(self, account_id, account_name, role_name):
        """
        Obtém dados de custo para uma conta específica.
        
        Args:
            account_id: ID da conta AWS
            account_name: Nome da conta AWS
            role_name: Nome da função a assumir
            
        Returns:
            dict: Dicionário com os dados de custo
        """
        try:
            account_session = self.sso_auth.get_credentials(
                self.access_token, account_id, role_name
            )
            
            if not account_session:
                return None
                
            cost_explorer = account_session.client('ce')
            
            # Define o período dos últimos 3 meses
            end_date = datetime.datetime.now().date().replace(day=1)
            start_date = (end_date - relativedelta(months=3)).replace(day=1)
            
            start_str = start_date.strftime('%Y-%m-%d')
            end_str = end_date.strftime('%Y-%m-%d')
            
            # Obter custo total
            total_cost_response = cost_explorer.get_cost_and_usage(
                TimePeriod={
                    'Start': start_str,
                    'End': end_str
                },
                Granularity='MONTHLY',
                Metrics=['UnblendedCost']
            )
            
            # Processar resultados do custo total
            months = []
            month_start_dates = []
            total_costs = []
            
            # Extrair custos totais por mês
            for result in total_cost_response['ResultsByTime']:
                period = result['TimePeriod']
                start_date = period['Start']
                month_start_dates.append(start_date)
                month_obj = datetime.datetime.strptime(start_date, '%Y-%m-%d')
                month_name = month_obj.strftime('%b/%Y')
                months.append(month_name)
                
                amount = float(result['Total']['UnblendedCost']['Amount'])
                total_costs.append(amount)
            
            # Obter os top 10 serviços por custo para esta conta específica
            top_services_response = cost_explorer.get_cost_and_usage(
                TimePeriod={
                    'Start': start_str,
                    'End': end_str
                },
                Granularity='MONTHLY',
                GroupBy=[
                    {
                        'Type': 'DIMENSION',
                        'Key': 'SERVICE'
                    }
                ],
                Metrics=['UnblendedCost']
            )
            
            # Processar resultados por serviço
            service_total_costs = {}
            
            # Calcular o custo total para cada serviço em todos os meses
            for result in top_services_response['ResultsByTime']:
                for group in result['Groups']:
                    service_name = group['Keys'][0]
                    amount = float(group['Metrics']['UnblendedCost']['Amount'])
                    
                    if service_name in service_total_costs:
                        service_total_costs[service_name] += amount
                    else:
                        service_total_costs[service_name] = amount
            
            # Ordenar serviços por custo total e pegar os top 10 para esta conta
            sorted_services = sorted(service_total_costs.items(), key=lambda x: x[1], reverse=True)
            top_10_services = sorted_services[:10]
            
            # Criar mapeamento de nomes AWS para nomes amigáveis
            service_name_mapping = {}
            for service_name, _ in top_10_services:
                # Extrair nome mais curto para serviços da AWS
                if service_name.startswith('Amazon '):
                    short_name = service_name.replace('Amazon ', '')
                elif service_name.startswith('AWS '):
                    short_name = service_name.replace('AWS ', '')
                else:
                    short_name = service_name
                
                # Abreviar mais se o nome ainda for muito longo
                if len(short_name) > 20:
                    words = short_name.split()
                    if len(words) > 1:
                        short_name = ''.join([w[0] for w in words if w[0].isupper()])
                        if len(short_name) < 2:  # Se não tiver iniciais suficientes
                            short_name = short_name + words[-1]
                
                service_name_mapping[service_name] = short_name
            
            # Mapear serviços de top 10 para análise mensal
            top_10_service_keys = [service_name_mapping[name] for name, _ in top_10_services]
            top_10_service_names = [name for name, _ in top_10_services]
            
            # Dicionário para armazenar custos de cada serviço por mês
            service_costs = {}
            service_percentages = {}
            
            # Inicializar dicionários para serviços
            for i, service_key in enumerate(top_10_service_keys):
                service_costs[service_key] = [0.0] * len(months)
                service_percentages[service_key] = [0.0] * len(months)
            
            # Para cada mês, obter os custos por serviço
            for month_idx, start_date in enumerate(month_start_dates):
                # Obter detalhes do mês
                month_response = cost_explorer.get_cost_and_usage(
                    TimePeriod={
                        'Start': start_date,
                        'End': (datetime.datetime.strptime(start_date, '%Y-%m-%d') + 
                                relativedelta(months=1)).strftime('%Y-%m-%d')
                    },
                    Granularity='MONTHLY',
                    GroupBy=[
                        {
                            'Type': 'DIMENSION',
                            'Key': 'SERVICE'
                        }
                    ],
                    Metrics=['UnblendedCost']
                )
                
                # Processar serviços para este mês
                for group in month_response['ResultsByTime'][0]['Groups']:
                    service_name = group['Keys'][0]
                    amount = float(group['Metrics']['UnblendedCost']['Amount'])
                    
                    # Se o serviço está nos top 10, armazenar seu custo
                    if service_name in top_10_service_names:
                        service_idx = top_10_service_names.index(service_name)
                        service_key = top_10_service_keys[service_idx]
                        service_costs[service_key][month_idx] = amount
                        
                        # Calcular percentual se houver custo total
                        if total_costs[month_idx] > 0:
                            percentage = (amount / total_costs[month_idx]) * 100
                        else:
                            percentage = 0.0
                        service_percentages[service_key][month_idx] = percentage
            
            # Calcular totais para os 3 meses
            total_3_months = sum(total_costs)
            
            service_totals = {}
            service_total_percentages = {}
            
            for i, service_key in enumerate(top_10_service_keys):
                service_total = sum(service_costs[service_key])
                service_totals[service_key] = service_total
                
                if total_3_months > 0:
                    service_total_percentage = (service_total / total_3_months) * 100
                else:
                    service_total_percentage = 0.0
                    
                service_total_percentages[service_key] = service_total_percentage
            
            return {
                'account_id': account_id,
                'account_name': account_name,
                'role_name': role_name,
                'months': months,
                'month_start_dates': month_start_dates,
                'total_costs': total_costs,
                'service_costs': service_costs,
                'service_percentages': service_percentages,
                'total_3_months': total_3_months,
                'service_totals': service_totals,
                'service_total_percentages': service_total_percentages,
                'service_name_mapping': service_name_mapping,
                'top_10_service_keys': top_10_service_keys,
                'top_10_service_names': top_10_service_names
            }
            
        except Exception as e:
            print(f"Erro ao obter dados de custo para a conta {account_id}: {str(e)}")
            return None
    
    def extract_all_accounts_cost(self, account_filter=None, max_accounts=None, role_name_preference=None):
        """
        Extrai dados de custo para todas as contas.
        
        Args:
            account_filter: Filtro para nomes/IDs de contas (regex)
            max_accounts: Número máximo de contas a processar
            role_name_preference: Lista de nomes de funções por ordem de preferência
        """
        accounts = self.sso_auth.list_available_accounts(self.access_token)
        
        # Aplicar filtro se especificado
        if account_filter:
            try:
                filter_regex = re.compile(account_filter, re.IGNORECASE)
                accounts = [account for account in accounts 
                           if filter_regex.search(account['accountName']) or 
                           filter_regex.search(account['accountId'])]
                print(f"Filtro aplicado. {len(accounts)} contas correspondem ao filtro '{account_filter}'.")
            except re.error as e:
                print(f"Erro na expressão regular: {str(e)}. Ignorando filtro.")
        
        # Limitar número de contas se especificado
        if max_accounts and max_accounts > 0 and max_accounts < len(accounts):
            accounts = accounts[:max_accounts]
            print(f"Limitando a {max_accounts} contas para processamento.")
            
        total_accounts = len(accounts)
        
        if total_accounts == 0:
            print("Nenhuma conta encontrada para extrair dados. Verifique seu filtro ou permissões SSO.")
            return
            
        print(f"Encontradas {total_accounts} contas. Iniciando extração de custos...")
        
        for i, account in enumerate(accounts, 1):
            account_id = account['accountId']
            account_name = account['accountName']
            print(f"Processando conta {i}/{total_accounts}: {account_name} ({account_id})")
            
            # Obter funções disponíveis para a conta
            roles = self.sso_auth.get_account_roles(self.access_token, account_id)
            
            if not roles:
                print(f"⚠️ Nenhuma função disponível para a conta {account_name}")
                continue
                
            # Selecionar função apropriada
            selected_role = None
            
            # Se tiver uma lista de preferências, use-a
            if role_name_preference:
                for preferred_role in role_name_preference:
                    for role in roles:
                        if role['roleName'] == preferred_role:
                            selected_role = preferred_role
                            break
                    if selected_role:
                        break
            
            # Se não tiver preferência ou nenhuma função preferida estiver disponível,
            # use a primeira função com "Admin" ou a primeira disponível
            if not selected_role:
                for role in roles:
                    if 'Admin' in role['roleName'] or 'admin' in role['roleName']:
                        selected_role = role['roleName']
                        break
                        
                if not selected_role and roles:
                    selected_role = roles[0]['roleName']
            
            if not selected_role:
                print(f"⚠️ Não foi possível selecionar uma função para a conta {account_name}")
                continue
                
            print(f"Usando função: {selected_role}")
            
            cost_data = self.get_cost_data(account_id, account_name, selected_role)
            if cost_data:
                self.account_data.append(cost_data)
                print(f"✅ Dados extraídos com sucesso para {account_name}")
            else:
                print(f"❌ Falha ao extrair dados para {account_name}")
        
        print(f"Extração concluída para {len(self.account_data)} de {total_accounts} contas.")
    
    def _prepare_excel_data_per_account(self):
        """
        Prepara os dados para o relatório Excel, analisando cada conta individualmente.
        
        Returns:
            tuple: (summary_df_list, monthly_dfs_list, all_monthly_df) - DataFrames para as diferentes abas do Excel
        """
        if not self.account_data:
            print("Sem dados para gerar relatório.")
            return None, None, None
            
        # Lista para armazenar um resumo para cada conta
        summary_df_list = []
        
        # Dicionário para armazenar os DataFrames mensais para cada conta
        monthly_dfs_list = []
        
        # Dados para o DataFrame de todos os meses
        all_monthly_data = []
        
        # Obter todos os meses únicos em ordem cronológica
        all_months = []
        for account in self.account_data:
            for month in account['months']:
                if month not in all_months:
                    all_months.append(month)
        
        # Ordenar meses (formato 'Mmm/YYYY')
        all_months.sort(key=lambda x: (int(x.split('/')[1]), ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez'].index(x.split('/')[0])))
        
        # Processar cada conta individualmente
        for account in self.account_data:
            account_id = account['account_id']
            account_name = account['account_name']
            
            # Criar DataFrame de resumo para esta conta
            account_summary_data = {
                'ID da Conta': account_id,
                'Nome da Conta': account_name,
                'Função': account['role_name'],
                'Custo Total (USD)': account['total_3_months']
            }
            
            # Adicionar dados para cada um dos top 10 serviços desta conta
            for service_key in account['top_10_service_keys']:
                account_summary_data[f'{service_key} (USD)'] = account['service_totals'][service_key]
                account_summary_data[f'%'] = account['service_total_percentages'][service_key]
            
            # Criar DataFrame de resumo para esta conta
            summary_df = pd.DataFrame([account_summary_data])
            
            # Adicionar à lista de resumos
            summary_df_list.append({
                'account_id': account_id,
                'account_name': account_name,
                'df': summary_df
            })
            
            # Criar DataFrames mensais para esta conta
            account_monthly_dfs = {}
            
            for month in all_months:
                if month in account['months']:
                    i = account['months'].index(month)
                    
                    monthly_data = {
                        'ID da Conta': account_id,
                        'Nome da Conta': account_name,
                        'Custo Total (USD)': account['total_costs'][i]
                    }
                    
                    # Adicionar dados para cada um dos top 10 serviços desta conta
                    for service_key in account['top_10_service_keys']:
                        monthly_data[f'{service_key} (USD)'] = account['service_costs'][service_key][i]
                        monthly_data[f'%'] = account['service_percentages'][service_key][i]
                    
                    # Adicionar este mês ao DataFrame mensal da conta
                    account_monthly_dfs[month] = pd.DataFrame([monthly_data])
                    
                    # Adicionar linha para o DataFrame completo de todos os meses
                    all_monthly_row = {
                        'ID da Conta': account_id,
                        'Nome da Conta': account_name,
                        'Mês': month,
                        'Custo Total (USD)': account['total_costs'][i]
                    }
                    
                    # Adicionar dados de serviços para o DataFrame de todos os meses
                    for service_key in account['top_10_service_keys']:
                        all_monthly_row[f'{service_key} (USD)'] = account['service_costs'][service_key][i]
                        all_monthly_row[f'{service_key} %'] = account['service_percentages'][service_key][i]
                    
                    all_monthly_data.append(all_monthly_row)
            
            # Adicionar à lista de DataFrames mensais
            monthly_dfs_list.append({
                'account_id': account_id,
                'account_name': account_name,
                'monthly_dfs': account_monthly_dfs
            })
        
        # Criar DataFrame para todos os meses
        all_monthly_df = pd.DataFrame(all_monthly_data)
        
        # Ordenar por conta e mês
        if not all_monthly_df.empty:
            all_monthly_df = all_monthly_df.sort_values(['Nome da Conta', 'Mês'])
        
        return summary_df_list, monthly_dfs_list, all_monthly_df

    def generate_excel_report(self, output_path="aws_cost_report.xlsx"):
        """
        Gera um relatório Excel com os dados de custo extraídos.
        
        Args:
            output_path: Caminho para salvar o arquivo Excel
        """
        summary_df_list, monthly_dfs_list, all_monthly_df = self._prepare_excel_data_per_account()
        
        if not summary_df_list:
            return
        
        # Criar arquivo Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Adicionar uma aba de resumo para cada conta
            for account_summary in summary_df_list:
                account_name = account_summary['account_name']
                account_id = account_summary['account_id']
                df = account_summary['df']
                
                # Limitar o nome da aba a 31 caracteres (limite do Excel)
                sheet_name = f"Resumo_{account_name}"
                if len(sheet_name) > 31:
                    sheet_name = f"Resumo_{account_id}"
                    if len(sheet_name) > 31:
                        sheet_name = f"Resumo_{account_id[-8:]}"
                
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Formatar a planilha
                self._format_worksheet(
                    worksheet=writer.sheets[sheet_name],
                    df=df,
                    border=Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    ),
                    total_col_idx=df.columns.get_loc('Custo Total (USD)'),
                    service_keys=[col.split(' ')[0] for col in df.columns if ' (USD)' in col]
                )
            
            # Adicionar abas mensais para cada conta
            for account_monthly in monthly_dfs_list:
                account_name = account_monthly['account_name']
                account_id = account_monthly['account_id']
                account_monthly_dfs = account_monthly['monthly_dfs']
                
                # Adicionar uma aba para cada mês desta conta
                for month, df in account_monthly_dfs.items():
                    # Limitar o nome da aba a 31 caracteres (limite do Excel)
                    safe_month = month.replace('/', '-')
                    sheet_name = f"{account_name}_{safe_month}"
                    if len(sheet_name) > 31:
                        sheet_name = f"{account_id}_{safe_month}"
                        if len(sheet_name) > 31:
                            sheet_name = f"{account_id[-8:]}_{safe_month}"
                    
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Formatar a planilha
                    self._format_worksheet(
                        worksheet=writer.sheets[sheet_name],
                        df=df,
                        border=Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        ),
                        total_col_idx=df.columns.get_loc('Custo Total (USD)'),
                        service_keys=[col.split(' ')[0] for col in df.columns if ' (USD)' in col]
                    )
            
            # Adicionar a aba de todos os meses
            if not all_monthly_df.empty:
                all_monthly_df.to_excel(writer, sheet_name='Todos os Meses', index=False)
                
                # Formatar a planilha
                self._format_worksheet(
                    worksheet=writer.sheets['Todos os Meses'],
                    df=all_monthly_df,
                    border=Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    ),
                    total_col_idx=all_monthly_df.columns.get_loc('Custo Total (USD)'),
                    service_keys=[col.split(' ')[0] for col in all_monthly_df.columns if ' (USD)' in col]
                )
            
            # Adicionar filtros a todas as planilhas
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                workbook[sheet_name].auto_filter.ref = workbook[sheet_name].dimensions
        
        print(f"Relatório Excel gerado com sucesso: {output_path}")


    def _format_worksheet(self, worksheet, df, border, total_col_idx, service_keys):
        """
        Formata uma planilha Excel.
        
        Args:
            worksheet: Planilha a ser formatada
            df: DataFrame com os dados
            border: Estilo de borda
            total_col_idx: Índice da coluna de custo total
            service_keys: Lista de chaves dos serviços
        """
        # Definir largura das colunas
        for idx, col in enumerate(df.columns):
            column_width = max(len(col) + 2, df[col].astype(str).map(len).max() + 2)
            column_width = min(column_width, 40)  # Limitar largura máxima
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = column_width
        
        # Formatar todas as células com borda
        for row in range(1, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border
        
        # Formatar células numéricas
        for row in range(2, len(df) + 2):
            # Formatar custo total
            total_col = total_col_idx + 1  # Ajuste para indexação base-1 do openpyxl
            worksheet.cell(row=row, column=total_col).number_format = '$#,##0.00'
            
            # Formatar colunas de serviços
            for col_idx, col_name in enumerate(df.columns, 1):
                if ' (USD)' in col_name:
                    worksheet.cell(row=row, column=col_idx).number_format = '$#,##0.00'
                elif '% ' in col_name or col_name.endswith(' %'):
                    percent_cell = worksheet.cell(row=row, column=col_idx)
                    percent_cell.number_format = '0.00%'
                    
                    # Destacar percentuais altos (> 10%)
                    percentage = percent_cell.value
                    if percentage and percentage > 10:
                        percent_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Formatar cabeçalhos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Congelar primeira linha
        worksheet.freeze_panes = 'A2'
        
        # Adicionar regras de formatação condicional
        # Destacar valores maiores em tons mais escuros de azul para colunas de custo
        cost_columns = []
        for col_idx, col_name in enumerate(df.columns, 1):
            if ' (USD)' in col_name:
                cost_columns.append(col_idx)
        
        for col_idx in cost_columns:
            col_letter = get_column_letter(col_idx)
            max_row = len(df) + 1
            worksheet.conditional_formatting.add(
                f'{col_letter}2:{col_letter}{max_row}',
                ColorScaleRule(
                    start_type='min',
                    start_color='EBEDF0',
                    end_type='max',
                    end_color='4472C4'
                )
            )
    
    def generate_html_report(self, output_path="aws_cost_report.html"):
        """
        Gera um relatório HTML com os dados de custo extraídos.
        
        Args:
            output_path: Caminho para salvar o arquivo HTML
        """
        summary_df_list, monthly_dfs_list, all_monthly_df = self._prepare_excel_data_per_account()
        
        if not summary_df_list:
            return
        
        # Combinar os resumos de todas as contas para a visualização principal
        combined_summary_data = []
        for account_summary in summary_df_list:
            df = account_summary['df']
            for _, row in df.iterrows():
                combined_summary_data.append(row.to_dict())
        
        # Criar DataFrame combinado de resumo
        if combined_summary_data:
            combined_summary_df = pd.DataFrame(combined_summary_data)
        else:
            combined_summary_df = pd.DataFrame()
        
        # Formatar os dados para exibição HTML
        all_monthly_df_display = all_monthly_df.copy() if not all_monthly_df.empty else pd.DataFrame()
        
        # Dicionário para versões formatadas dos DataFrames
        account_summary_display = {}
        account_monthly_display = {}
        
        # Formatar o DataFrame de todos os meses
        if not all_monthly_df_display.empty:
            all_monthly_df_display['Custo Total (USD)'] = all_monthly_df_display['Custo Total (USD)'].map('${:,.2f}'.format)
            
            # Formatar colunas de serviços
            for col in all_monthly_df_display.columns:
                if ' (USD)' in col:
                    all_monthly_df_display[col] = all_monthly_df_display[col].apply(
                        lambda x: f"${x:,.2f}" if isinstance(x, (int, float)) else x
                    )
                elif '%' in col:  # incluindo todas as colunas que contêm '%'
                    all_monthly_df_display[col] = all_monthly_df_display[col].apply(
                        lambda x: f"{x:.2f}%" if isinstance(x, (int, float)) else x
                    )
        
        # Formatar resumos por conta
        for account_summary in summary_df_list:
            account_id = account_summary['account_id']
            df = account_summary['df'].copy()
            
            # Formatar coluna de custo total
            df['Custo Total (USD)'] = df['Custo Total (USD)'].map('${:,.2f}'.format)
            
            # Formatar colunas de serviços
            for col in df.columns:
                if ' (USD)' in col:
                    df[col] = df[col].apply(
                        lambda x: f"${x:,.2f}" if isinstance(x, (int, float)) else x
                    )
                elif '%' in col:  # Para capturar qualquer coluna que contenha '%'
                    df[col] = df[col].apply(
                        lambda x: f"{x:.2f}%" if isinstance(x, (int, float)) else x
                    )
            
            account_summary_display[account_id] = df
        
        # Formatar DataFrames mensais por conta
        for account_monthly in monthly_dfs_list:
            account_id = account_monthly['account_id']
            monthly_dfs = account_monthly['monthly_dfs']
            
            account_monthly_display[account_id] = {}
            
            for month, df in monthly_dfs.items():
                df_display = df.copy()
                
                # Formatar coluna de custo total
                df_display['Custo Total (USD)'] = df_display['Custo Total (USD)'].map('${:,.2f}'.format)
                
                # Formatar colunas de serviços
                for col in df_display.columns:
                    if ' (USD)' in col:
                        df_display[col] = df_display[col].apply(
                            lambda x: f"${x:,.2f}" if isinstance(x, (int, float)) else x
                        )
                    elif '%' in col:  # Para capturar qualquer coluna que contenha '%'
                        df_display[col] = df_display[col].apply(
                            lambda x: f"{x:.2f}%" if isinstance(x, (int, float)) else x
                        )
                
                account_monthly_display[account_id][month] = df_display
        
        # Obter todos os meses únicos em ordem cronológica
        all_months = []
        for account_monthly in monthly_dfs_list:
            monthly_dfs = account_monthly['monthly_dfs']
            for month in monthly_dfs.keys():
                if month not in all_months:
                    all_months.append(month)
        
        # Ordenar meses
        all_months.sort(key=lambda x: (int(x.split('/')[1]), ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez'].index(x.split('/')[0])))
        
        # Gerar cabeçalhos de tabela para o resumo combinado
        combined_headers = ''
        if not combined_summary_df.empty:
            combined_headers = ''.join([f'<th data-sort="{col}" class="sortable">{col}</th>' for col in combined_summary_df.columns])
        
        # Gerar linhas da tabela de resumo combinado
        combined_rows = []
        if not combined_summary_df.empty:
            for _, row in combined_summary_df.iterrows():
                cells = []
                for col, value in row.items():
                    cell_class = "cost-cell" if "USD" in col or "%" in col else ""
                    cells.append(f'<td class="{cell_class}">{value}</td>')
                combined_rows.append('<tr>' + ''.join(cells) + '</tr>')
        combined_table_body = ''.join(combined_rows)
        
        # Gerar cabeçalhos para a tabela de todos os meses
        all_monthly_headers = ''
        if not all_monthly_df_display.empty:
            all_monthly_headers = ''.join([f'<th data-sort="{col}" class="sortable">{col}</th>' for col in all_monthly_df_display.columns])
        
        # Gerar linhas da tabela de todos os meses
        all_monthly_rows = []
        if not all_monthly_df_display.empty:
            for _, row in all_monthly_df_display.iterrows():
                cells = []
                for col, value in row.items():
                    cell_class = "cost-cell" if "USD" in col or "%" in col else ""
                    cells.append(f'<td class="{cell_class}">{value}</td>')
                all_monthly_rows.append('<tr>' + ''.join(cells) + '</tr>')
        all_monthly_table_body = ''.join(all_monthly_rows)
        
        # Gerar abas para as contas
        account_tabs = []
        for i, account_summary in enumerate(summary_df_list):
            account_name = account_summary['account_name']
            account_id = account_summary['account_id']
            active_class = "" if i > 0 else "active"
            account_tabs.append(f'<div class="tab account-tab {active_class}" id="tab-account-{account_id}" onclick="showAccountTab(\'{account_id}\')">{account_name}</div>')
        account_tabs_html = ''.join(account_tabs)
        
        # Gerar conteúdo das abas de contas (resumo + meses)
        account_contents = []
        for i, account_summary in enumerate(summary_df_list):
            account_id = account_summary['account_id']
            account_name = account_summary['account_name']
            
            display_style = "block" if i == 0 else "none"
            
            # Gerar cabeçalhos de tabela para esta conta
            summary_headers = ''.join([f'<th data-sort="{col}" class="sortable">{col}</th>' for col in account_summary_display[account_id].columns])
            
            # Gerar linhas da tabela de resumo para esta conta
            summary_rows = []
            for _, row in account_summary_display[account_id].iterrows():
                cells = []
                for col, value in row.items():
                    cell_class = "cost-cell" if "USD" in col or "%" in col else ""
                    cells.append(f'<td class="{cell_class}">{value}</td>')
                summary_rows.append('<tr>' + ''.join(cells) + '</tr>')
            summary_table_body = ''.join(summary_rows)
            
            # Gerar abas para os meses desta conta
            account_month_tabs = []
            account_months = list(account_monthly_display[account_id].keys()) if account_id in account_monthly_display else []
            
            for j, month in enumerate(account_months):
                month_active_class = "" if j > 0 else "active"
                month_id = month.replace('/', '-')
                account_month_tabs.append(f'<div class="tab month-tab {month_active_class}" id="tab-{account_id}-{month_id}" onclick="showAccountMonthTab(\'{account_id}\', \'{month_id}\')">{month}</div>')
            account_month_tabs_html = ''.join(account_month_tabs)
            
            # Gerar conteúdo das abas mensais para esta conta
            account_month_contents = []
            for j, month in enumerate(account_months):
                month_id = month.replace('/', '-')
                month_display_style = "block" if j == 0 else "none"
                
                # Gerar cabeçalhos para este mês e conta
                monthly_headers = ''.join([f'<th data-sort="{col}" class="sortable">{col}</th>' for col in account_monthly_display[account_id][month].columns])
                
                # Gerar linhas da tabela para este mês e conta
                monthly_rows = []
                for _, row in account_monthly_display[account_id][month].iterrows():
                    cells = []
                    for col, value in row.items():
                        cell_class = "cost-cell" if "USD" in col or "%" in col else ""
                        cells.append(f'<td class="{cell_class}">{value}</td>')
                    monthly_rows.append('<tr>' + ''.join(cells) + '</tr>')
                monthly_table_body = ''.join(monthly_rows)
                
                account_month_contents.append(f'''
                <div id="month-{account_id}-{month_id}" class="month-content" style="display: {month_display_style};">
                    <div style="overflow-x: auto;">
                        <table id="table-{account_id}-{month_id}" class="sortable-table">
                            <thead>
                                <tr>
                                    {monthly_headers}
                                </tr>
                            </thead>
                            <tbody>
                                {monthly_table_body}
                            </tbody>
                        </table>
                    </div>
                </div>''')
            account_month_contents_html = ''.join(account_month_contents)
            
            # Legenda dos serviços para esta conta
            service_keys = [col.split(' ')[0] for col in account_summary_display[account_id].columns if ' (USD)' in col]
            service_legend = []
            for service_key in service_keys:
                color = f'#{abs(hash(service_key)) % 0xffffff:06x}'
                service_legend.append(f'<div class="service-item"><div class="service-color" style="background-color: {color}"></div>{service_key}</div>')
            service_legend_html = ''.join(service_legend)
            
            # Montando o conteúdo da aba desta conta
            account_contents.append(f'''
            <div id="account-{account_id}" class="account-content" style="display: {display_style};">
                <h2>Conta: {account_name} ({account_id})</h2>
                
                <input type="text" id="search-account-{account_id}" class="search" placeholder="Buscar..." onkeyup="filterAccountTable('{account_id}')">
                
                <div class="service-legend">
                    <h3>Serviços para esta conta:</h3>
                    <div style="display: flex; flex-wrap: wrap; gap: 10px; margin-top: 10px;">
                        {service_legend_html}
                    </div>
                </div>
                
                <div class="tabs account-view-tabs">
                    <div class="tab active" onclick="showAccountViewTab('{account_id}', 'resumo')">Resumo</div>
                    <div class="tab" onclick="showAccountViewTab('{account_id}', 'mensal')">Detalhes por Mês</div>
                </div>
                
                <div id="account-resumo-{account_id}" class="account-view-content" style="display: block;">
                    <div style="overflow-x: auto;">
                        <table id="resumo-table-{account_id}" class="sortable-table">
                            <thead>
                                <tr>
                                    {summary_headers}
                                </tr>
                            </thead>
                            <tbody>
                                {summary_table_body}
                            </tbody>
                        </table>
                    </div>
                </div>
                
                <div id="account-mensal-{account_id}" class="account-view-content" style="display: none;">
                    <div class="month-tabs">
                        {account_month_tabs_html}
                    </div>
                    {account_month_contents_html}
                </div>
            </div>''')
        account_contents_html = ''.join(account_contents)
        
        # Gerar timestamp
        timestamp = datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        
        # Combinar tudo em um HTML
        html_content = f"""<!DOCTYPE html>
    <html lang="pt-br">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Relatório de Custos AWS</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                margin: 0;
                padding: 20px;
                color: #333;
            }}
            h1, h2, h3 {{
                color: #0066cc;
            }}
            .container {{
                max-width: 1200px;
                margin: 0 auto;
            }}
            table {{
                border-collapse: collapse;
                width: 100%;
                margin-bottom: 30px;
                font-size: 14px;
            }}
            th, td {{
                text-align: left;
                padding: 10px;
                border: 1px solid #ddd;
            }}
            th {{
                background-color: #0066cc;
                color: white;
                position: sticky;
                top: 0;
                z-index: 10;
                cursor: pointer;
            }}
            th.sortable:hover {{
                background-color: #0055aa;
            }}
            th.sortable:after {{
                content: "\\00a0\\00a0\\00a0"; /* Espaço para o ícone de ordenação */
            }}
            th.sort-asc:after {{
                content: "\\2191"; /* Seta para cima */
            }}
            th.sort-desc:after {{
                content: "\\2193"; /* Seta para baixo */
            }}
            tr:nth-child(even) {{
                background-color: #f5f5f5;
            }}
            tr:hover {{
                background-color: #e9f1fa;
            }}
            .high-percentage {{
                background-color: #FFEB9C;
            }}
            .timestamp {{
                font-size: 0.8em;
                color: #666;
                margin-bottom: 20px;
            }}
            .summary {{
                background-color: #f0f7ff;
                padding: 15px;
                border-radius: 5px;
                margin-bottom: 20px;
            }}
            .tabs {{
                display: flex;
                margin-bottom: 20px;
                border-bottom: 1px solid #ccc;
            }}
            .tab {{
                padding: 10px 20px;
                background-color: #ddd;
                cursor: pointer;
                border: 1px solid #ccc;
                border-bottom: none;
                border-radius: 5px 5px 0 0;
                margin-right: 5px;
            }}
            .tab.active {{
                background-color: #0066cc;
                color: white;
            }}
            .tab-content {{
                display: none;
            }}
            .tab-content.active {{
                display: block;
            }}
            .month-tabs {{
                display: flex;
                flex-wrap: wrap;
                margin-bottom: 20px;
                border-bottom: 1px solid #ccc;
            }}
            .month-tab {{
                padding: 8px 16px;
                background-color: #eee;
                cursor: pointer;
                border: 1px solid #ddd;
                border-bottom: none;
                border-radius: 4px 4px 0 0;
                margin-right: 4px;
                margin-bottom: 0;
                font-size: 0.9em;
            }}
            .month-tab.active {{
                background-color: #4a86e8;
                color: white;
            }}
            .month-content {{
                margin-top: 20px;
            }}
            .search {{
                padding: 10px;
                margin-bottom: 20px;
                width: 100%;
                box-sizing: border-box;
                border: 1px solid #ddd;
                border-radius: 4px;
            }}
            .cost-cell {{
                text-align: right;
            }}
            .service-legend {{
                margin-bottom: 20px;
                display: flex;
                flex-wrap: wrap;
            }}
            .service-item {{
                margin-right: 20px;
                margin-bottom: 10px;
                display: flex;
                align-items: center;
            }}
            .service-color {{
                width: 20px;
                height: 20px;
                margin-right: 5px;
                border-radius: 3px;
            }}
            .dashboard {{
                display: flex;
                flex-wrap: wrap;
                gap: 20px;
                margin-bottom: 20px;
            }}
            .dashboard-card {{
                background-color: #fff;
                border: 1px solid #ddd;
                border-radius: 5px;
                padding: 15px;
                flex: 1;
                min-width: 200px;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            }}
            .card-title {{
                font-size: 0.9em;
                margin-bottom: 10px;
                color: #666;
            }}
            .card-value {{
                font-size: 1.8em;
                font-weight: bold;
                color: #0066cc;
            }}
            .account-tabs {{
                display: flex;
                flex-wrap: wrap;
                margin-bottom: 20px;
            }}
            .account-tab {{
                padding: 10px 15px;
                margin-right: 5px;
                margin-bottom: 5px;
                font-size: 0.9em;
            }}
            .account-content {{
                margin-top: 20px;
            }}
            .account-view-tabs {{
                margin-bottom: 10px;
            }}
            .account-view-content {{
                margin-bottom: 30px;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <h1>Relatório de Custos AWS</h1>
            <div class="timestamp">Gerado em: {timestamp}</div>
            
            <div class="summary">
                <h2>Resumo de Custos</h2>
                <p>Total de Contas Analisadas: {len(summary_df_list)}</p>
            </div>
            
            <div class="tabs">
                <div class="tab active" onclick="showTab('contas')">Por Conta</div>
                <div class="tab" onclick="showTab('resumo')">Resumo Geral</div>
                <div class="tab" onclick="showTab('todos-meses')">Todos os Meses</div>
            </div>
            
            <div id="contas" class="tab-content active">
                <div class="account-tabs">
                    {account_tabs_html}
                </div>
                
                {account_contents_html}
            </div>
            
            <div id="resumo" class="tab-content">
                <input type="text" id="searchResumo" class="search" placeholder="Buscar por conta..." onkeyup="filterTable('resumoTable', 'searchResumo')">
                
                <div style="overflow-x: auto;">
                    <table id="resumoTable" class="sortable-table">
                        <thead>
                            <tr>
                                {combined_headers}
                            </tr>
                        </thead>
                        <tbody>
                            {combined_table_body}
                        </tbody>
                    </table>
                </div>
            </div>
            
            <div id="todos-meses" class="tab-content">
                <input type="text" id="searchTodosMeses" class="search" placeholder="Buscar por conta ou mês..." onkeyup="filterTable('todosMesesTable', 'searchTodosMeses')">
                
                <div style="overflow-x: auto;">
                    <table id="todosMesesTable" class="sortable-table">
                        <thead>
                            <tr>
                                {all_monthly_headers}
                            </tr>
                        </thead>
                        <tbody>
                            {all_monthly_table_body}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
        
        <script>
            // Variáveis globais para rastrear estado
            var activeMainTab = 'contas';
            var activeAccountId = '{summary_df_list[0]["account_id"] if summary_df_list else ""}';
            var activeAccountViewTab = {{}}; // Para cada conta, qual visualização está ativa (resumo/mensal)
            var activeMonthTab = {{}}; // Para cada conta, qual mês está ativo
            
            // Inicializar estado
            document.addEventListener('DOMContentLoaded', function() {{
                // Inicializar estados para cada conta
                const accountContents = document.querySelectorAll('.account-content');
                for (let i = 0; i < accountContents.length; i++) {{
                    const accountId = accountContents[i].id.replace('account-', '');
                    activeAccountViewTab[accountId] = 'resumo';
                    
                    // Encontrar o primeiro mês para cada conta
                    const monthTabs = document.querySelectorAll('.month-tab[id^="tab-' + accountId + '"]');
                    if (monthTabs.length > 0) {{
                        const firstMonthId = monthTabs[0].id.split('-').slice(2).join('-');
                        activeMonthTab[accountId] = firstMonthId;
                    }}
                }}
                
                // Destacar células com percentuais altos
                highlightHighPercentages();
                
                // Inicializar a funcionalidade de ordenação para todas as tabelas
                initSortableTables();
            }});
            
            function showTab(tabId) {{
                // Ocultar todos os conteúdos
                var contents = document.getElementsByClassName('tab-content');
                for (var i = 0; i < contents.length; i++) {{
                    contents[i].classList.remove('active');
                    contents[i].style.display = 'none';
                }}
                
                // Atualizar abas principais
                var tabs = document.querySelector('.tabs').querySelectorAll('.tab');
                for (var i = 0; i < tabs.length; i++) {{
                    tabs[i].classList.remove('active');
                }}
                
                // Mostrar conteúdo selecionado
                document.getElementById(tabId).classList.add('active');
                document.getElementById(tabId).style.display = 'block';
                
                // Atualizar aba ativa
                event.currentTarget.classList.add('active');
                
                // Atualizar estado global
                activeMainTab = tabId;
            }}
            
            function showAccountTab(accountId) {{
                // Ocultar todos os conteúdos de conta
                var accountContents = document.getElementsByClassName('account-content');
                for (var i = 0; i < accountContents.length; i++) {{
                    accountContents[i].style.display = 'none';
                }}
                
                // Atualizar abas de conta
                var accountTabs = document.getElementsByClassName('account-tab');
                for (var i = 0; i < accountTabs.length; i++) {{
                    accountTabs[i].classList.remove('active');
                }}
                
                // Mostrar conteúdo da conta selecionada
                document.getElementById('account-' + accountId).style.display = 'block';
                
                // Atualizar aba ativa
                document.getElementById('tab-account-' + accountId).classList.add('active');
                
                // Atualizar estado global
                activeAccountId = accountId;
            }}
            
            function showAccountViewTab(accountId, viewTabId) {{
                // Ocultar todos os conteúdos de visualização
                var viewContents = document.getElementById('account-' + accountId).querySelectorAll('.account-view-content');
                for (var i = 0; i < viewContents.length; i++) {{
                    viewContents[i].style.display = 'none';
                }}
                
                // Atualizar abas de visualização
                var viewTabs = document.getElementById('account-' + accountId).querySelectorAll('.account-view-tabs .tab');
                for (var i = 0; i < viewTabs.length; i++) {{
                    viewTabs[i].classList.remove('active');
                }}
                
                // Mostrar conteúdo da visualização selecionada
                document.getElementById('account-' + viewTabId + '-' + accountId).style.display = 'block';
                
                // Atualizar aba ativa (clicar no segundo filho da account-view-tabs)
                event.currentTarget.classList.add('active');
                
                // Atualizar estado global
                activeAccountViewTab[accountId] = viewTabId;
            }}
            
            function showAccountMonthTab(accountId, monthId) {{
                // Ocultar todos os conteúdos de mês para esta conta
                var monthContents = document.getElementById('account-mensal-' + accountId).querySelectorAll('.month-content');
                for (var i = 0; i < monthContents.length; i++) {{
                    monthContents[i].style.display = 'none';
                }}
                
                // Atualizar abas de mês para esta conta
                var monthTabs = document.getElementById('account-mensal-' + accountId).querySelectorAll('.month-tab');
                for (var i = 0; i < monthTabs.length; i++) {{
                    monthTabs[i].classList.remove('active');
                }}
                
                // Mostrar conteúdo do mês selecionado
                document.getElementById('month-' + accountId + '-' + monthId).style.display = 'block';
                
                // Atualizar aba ativa
                document.getElementById('tab-' + accountId + '-' + monthId).classList.add('active');
                
                // Atualizar estado global
                activeMonthTab[accountId] = monthId;
            }}
            
            function filterTable(tableId, inputId) {{
                var input, filter, table, tr, td, i, j, txtValue, found;
                input = document.getElementById(inputId);
                filter = input.value.toUpperCase();
                table = document.getElementById(tableId);
                tr = table.getElementsByTagName("tr");
                
                for (i = 1; i < tr.length; i++) {{
                    found = false;
                    td = tr[i].getElementsByTagName("td");
                    
                    for (j = 0; j < 3; j++) {{
                        if (td[j]) {{
                            txtValue = td[j].textContent || td[j].innerText;
                            if (txtValue.toUpperCase().indexOf(filter) > -1) {{
                                found = true;
                                break;
                            }}
                        }}
                    }}
                    
                    if (found) {{
                        tr[i].style.display = "";
                    }} else {{
                        tr[i].style.display = "none";
                    }}
                }}
            }}
            
            function filterAccountTable(accountId) {{
                const tableId = 'resumo-table-' + accountId;
                const inputId = 'search-account-' + accountId;
                
                var input, filter, table, tr, td, i, j, txtValue, found;
                input = document.getElementById(inputId);
                filter = input.value.toUpperCase();
                table = document.getElementById(tableId);
                tr = table.getElementsByTagName("tr");
                
                for (i = 1; i < tr.length; i++) {{
                    found = false;
                    td = tr[i].getElementsByTagName("td");
                    
                    for (j = 0; j < td.length; j++) {{
                        if (td[j]) {{
                            txtValue = td[j].textContent || td[j].innerText;
                            if (txtValue.toUpperCase().indexOf(filter) > -1) {{
                                found = true;
                                break;
                            }}
                        }}
                    }}
                    
                    if (found) {{
                        tr[i].style.display = "";
                    }} else {{
                        tr[i].style.display = "none";
                    }}
                }}
            }}
            
            function highlightHighPercentages() {{
                const tables = document.querySelectorAll('table');
                tables.forEach(table => {{
                    const headerRow = table.querySelector('tr');
                    if (!headerRow) return;
                    
                    const headers = headerRow.querySelectorAll('th');
                    
                    for (let i = 0; i < headers.length; i++) {{
                        if (headers[i].textContent.includes('%')) {{
                            const colIndex = i;
                            
                            const rows = table.querySelectorAll('tbody tr');
                            for (let j = 0; j < rows.length; j++) {{
                                const cell = rows[j].querySelectorAll('td')[colIndex];
                                if (cell) {{
                                    const percentText = cell.textContent.trim();
                                    const percentValue = parseFloat(percentText);
                                    if (!isNaN(percentValue) && percentValue > 10) {{
                                        cell.classList.add('high-percentage');
                                    }}
                                }}
                            }}
                        }}
                    }}
                }});
            }}
            
            // Função para inicializar a ordenação em todas as tabelas
            function initSortableTables() {{
                const sortableTables = document.querySelectorAll('.sortable-table');
                
                sortableTables.forEach(table => {{
                    const headers = table.querySelectorAll('th.sortable');
                    
                    headers.forEach((header, index) => {{
                        header.addEventListener('click', function() {{
                            sortTable(table, index, this);
                        }});
                    }});
                }});
            }}
            
            // Função para ordenar uma tabela
            function sortTable(table, colIndex, header) {{
                const rows = Array.from(table.querySelectorAll('tbody tr'));
                const thead = table.querySelector('thead');
                const headers = thead.querySelectorAll('th');
                const isAsc = header.classList.contains('sort-asc');
                
                // Remover classes de ordenação de todos os cabeçalhos
                headers.forEach(h => {{
                    h.classList.remove('sort-asc', 'sort-desc');
                }});
                
                // Adicionar classe de ordenação ao cabeçalho atual
                header.classList.add(isAsc ? 'sort-desc' : 'sort-asc');
                
                // Ordenar as linhas
                rows.sort((a, b) => {{
                    const cellA = a.querySelectorAll('td')[colIndex].textContent.trim();
                    const cellB = b.querySelectorAll('td')[colIndex].textContent.trim();
                    
                    // Verificar se a célula contém um valor monetário (começa com $)
                    if (cellA.startsWith('$') && cellB.startsWith('$')) {{
                        // Remover $ e vírgulas, depois converter para número
                        const numA = parseFloat(cellA.replace(/[$,]/g, ''));
                        const numB = parseFloat(cellB.replace(/[$,]/g, ''));
                        return isAsc ? numB - numA : numA - numB;
                    }}
                    // Verificar se a célula contém uma porcentagem
                    else if (cellA.endsWith('%') && cellB.endsWith('%')) {{
                        const numA = parseFloat(cellA);
                        const numB = parseFloat(cellB);
                        return isAsc ? numB - numA : numA - numB;
                    }}
                    // Ordenação padrão como texto
                    else {{
                        return isAsc ? 
                            cellB.localeCompare(cellA, undefined, {{numeric: true, sensitivity: 'base'}}) :
                            cellA.localeCompare(cellB, undefined, {{numeric: true, sensitivity: 'base'}});
                    }}
                }});
                
                // Reconstruir a tabela com linhas ordenadas
                const tbody = table.querySelector('tbody');
                tbody.innerHTML = '';
                
                rows.forEach(row => {{
                    tbody.appendChild(row);
                }});
            }}
        </script>
    </body>
    </html>"""
            
        # Salvar o arquivo HTML
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"Relatório HTML gerado com sucesso: {output_path}")


def main():
    parser = argparse.ArgumentParser(description='Extrai dados de custo da AWS para múltiplas contas via SSO.')
    parser.add_argument('--start-url', type=str, required=True,
                        help='URL do portal SSO (ex: itaulzprod.awsapps.com/start)')
    parser.add_argument('--region', type=str, default='us-east-1',
                        help='Região AWS principal (padrão: us-east-1)')
    parser.add_argument('--sso-region', type=str,
                        help='Região do SSO se diferente da região principal')
    parser.add_argument('--preferred-roles', type=str, nargs='+',
                        help='Lista de nomes de funções preferidas, em ordem de prioridade')
    parser.add_argument('--format', choices=['excel', 'html', 'both'], default='both', 
                        help='Formato do relatório: excel, html, ou both (ambos)')
    parser.add_argument('--output-excel', type=str, default='aws_cost_report.xlsx',
                        help='Caminho para salvar o relatório Excel')
    parser.add_argument('--output-html', type=str, default='aws_cost_report.html',
                        help='Caminho para salvar o relatório HTML')
    parser.add_argument('--account-filter', type=str,
                        help='Filtrar contas por nome ou ID (aceita expressões regulares)')
    parser.add_argument('--max-accounts', type=int,
                        help='Número máximo de contas a processar (útil para testes)')
    
    args = parser.parse_args()
    
    # Inicializar autenticador SSO
    sso_auth = AWSSSOAuth(
        start_url=args.start_url,
        region=args.region,
        sso_region=args.sso_region
    )
    
    # Autenticar e obter token
    print("Iniciando autenticação AWS SSO...")
    access_token = sso_auth.authenticate()
    
    if not access_token:
        print("Falha na autenticação. Encerrando.")
        sys.exit(1)
    
    # Inicializar extrator de custos
    extractor = AWSCostExtractor(sso_auth, access_token)
    
    # Extrair custos de todas as contas
    print("Iniciando extração de custos da AWS...")
    extractor.extract_all_accounts_cost(
        account_filter=args.account_filter,
        max_accounts=args.max_accounts,
        role_name_preference=args.preferred_roles
    )
    
    # Gerar relatórios conforme solicitado
    if args.format in ['excel', 'both']:
        extractor.generate_excel_report(output_path=args.output_excel)
    
    if args.format in ['html', 'both']:
        extractor.generate_html_report(output_path=args.output_html)
    
    print("Processo concluído!")


if __name__ == "__main__":
    main()